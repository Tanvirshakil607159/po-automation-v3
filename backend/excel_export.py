"""Multi-sheet Excel export with consumption calculator columns.
   Supports two-level grouping: PO Number → Item Category."""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Style constants
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
CALC_HEADER_FILL = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _safe_sheet_name(name: str) -> str:
    """Make a string safe for use as an Excel sheet name."""
    invalid = r'[]:*?/\\'
    safe = "".join(c for c in name if c not in invalid)
    return safe[:31] if safe else "Sheet"


def create_export_workbook(
    grouped_data: dict,
    consumption_values: dict | None = None,
) -> bytes:
    """
    Create a multi-sheet Excel workbook from two-level grouped data.

    Sheets are named as "PO_Item" (e.g., "PO-1001_Thread").
    Falls back to flat categories if po_groups is absent.
    """
    wb = Workbook()
    wb.remove(wb.active)

    base_headers = grouped_data.get("headers", [])
    calc_headers = ["Consumption/Unit", "Wastage %", "Total Required Qty"]

    po_groups = grouped_data.get("po_groups", {})

    # If no po_groups, fall back to old flat structure
    if not po_groups:
        categories = grouped_data.get("categories", {})
        po_groups = {"All": {"categories": categories}}

    for po_name, po_data in po_groups.items():
        categories = po_data.get("categories", {})
        for cat_name, rows in categories.items():
            # Sheet name: "PO_Item" truncated to 31 chars
            sheet_label = f"{po_name}_{cat_name}"
            ws = wb.create_sheet(title=_safe_sheet_name(sheet_label))
            all_headers = base_headers + calc_headers

            # Write header row
            for col_idx, header in enumerate(all_headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = HEADER_FONT
                cell.fill = CALC_HEADER_FILL if header in calc_headers else HEADER_FILL
                cell.alignment = CENTER
                cell.border = THIN_BORDER

            # Write data rows
            grand_total = 0
            for row_idx, row_data in enumerate(rows):
                excel_row = row_idx + 2

                for col_idx, header in enumerate(base_headers, 1):
                    cell = ws.cell(row=excel_row, column=col_idx, value=row_data.get(header, ""))
                    cell.border = THIN_BORDER
                    cell.alignment = CENTER

                # Consumption values key: "po_name::cat_name::row_idx"
                cons_data = {}
                if consumption_values:
                    po_cons = consumption_values.get(po_name, {})
                    cat_cons = po_cons.get(cat_name, {}) if isinstance(po_cons, dict) else {}
                    cons_data = cat_cons.get(str(row_idx), {}) if isinstance(cat_cons, dict) else {}

                consumption = float(cons_data.get("consumption", 0))
                wastage = float(cons_data.get("wastage", 5))

                order_qty = 0
                for h in base_headers:
                    h_lower = h.lower()
                    if any(kw in h_lower for kw in ["qty", "quantity", "order qty"]):
                        try:
                            order_qty = float(str(row_data.get(h, "0")).replace(",", ""))
                        except (ValueError, TypeError):
                            order_qty = 0
                        break

                total_required = (order_qty * consumption) * (1 + wastage / 100)
                grand_total += total_required

                cons_col = len(base_headers) + 1
                cell_cons = ws.cell(row=excel_row, column=cons_col, value=consumption)
                cell_cons.fill = INPUT_FILL
                cell_cons.border = THIN_BORDER
                cell_cons.alignment = CENTER

                cell_waste = ws.cell(row=excel_row, column=cons_col + 1, value=wastage)
                cell_waste.fill = INPUT_FILL
                cell_waste.border = THIN_BORDER
                cell_waste.alignment = CENTER

                cell_total = ws.cell(row=excel_row, column=cons_col + 2, value=round(total_required, 2))
                cell_total.border = THIN_BORDER
                cell_total.alignment = CENTER

            # Grand total row
            total_row = len(rows) + 2
            ws.cell(row=total_row, column=len(base_headers), value="GRAND TOTAL").font = TOTAL_FONT
            ws.cell(row=total_row, column=len(base_headers)).fill = TOTAL_FILL
            ws.cell(row=total_row, column=len(base_headers)).border = THIN_BORDER

            total_cell = ws.cell(row=total_row, column=len(all_headers), value=round(grand_total, 2))
            total_cell.font = TOTAL_FONT
            total_cell.fill = TOTAL_FILL
            total_cell.border = THIN_BORDER
            total_cell.alignment = CENTER

            # Auto-width
            for col_idx in range(1, len(all_headers) + 1):
                col_letter = get_column_letter(col_idx)
                max_len = max(
                    len(str(ws.cell(row=r, column=col_idx).value or ""))
                    for r in range(1, total_row + 1)
                )
                ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
